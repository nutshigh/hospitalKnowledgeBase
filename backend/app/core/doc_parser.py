from pathlib import Path
from typing import List


class TextChunk:
    def __init__(self, text: str, title: str = "", chunk_index: int = 0):
        self.text = text
        self.title = title
        self.chunk_index = chunk_index


def parse_file(file_path: str, filename: str) -> List[TextChunk]:
    ext = Path(filename).suffix.lower()
    if ext == ".pdf":
        return _parse_pdf(file_path)
    elif ext in (".docx", ".doc"):
        return _parse_docx(file_path)
    elif ext in (".xlsx", ".xls"):
        return _parse_excel(file_path)
    elif ext in (".txt", ".md"):
        return _parse_text(file_path)
    else:
        raise ValueError(f"Unsupported file format: {ext}")


def _parse_pdf(file_path: str) -> List[TextChunk]:
    import fitz
    doc = fitz.open(file_path)
    full_text = ""
    for page in doc:
        text = page.get_text()
        if text.strip():
            full_text += text + "\n"
    doc.close()
    return _split_text(full_text, filename=Path(file_path).name)


def _parse_docx(file_path: str) -> List[TextChunk]:
    from docx import Document
    doc = Document(file_path)
    full_text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    return _split_text(full_text, filename=Path(file_path).name)


def _parse_excel(file_path: str) -> List[TextChunk]:
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True)
    chunks = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            row_text = " | ".join(str(c) if c is not None else "" for c in row)
            if row_text.strip().replace("|", "").strip():
                rows.append(row_text)
        if rows:
            text = f"Sheet: {sheet_name}\n" + "\n".join(rows)
            chunks.extend(_split_text(text, title=sheet_name))
    wb.close()
    return chunks


def _parse_text(file_path: str) -> List[TextChunk]:
    with open(file_path, "r", encoding="utf-8") as f:
        text = f.read()
    return _split_text(text, filename=Path(file_path).name)


def _split_text(
    text: str,
    filename: str = "",
    title: str = "",
    chunk_size: int = 500,
    overlap: int = 50,
) -> List[TextChunk]:
    chunks = []
    text = text.strip()
    if not text:
        return chunks

    paragraphs = text.split("\n")
    current = ""
    for para in paragraphs:
        if len(current) + len(para) > chunk_size and current:
            chunk_title = title or filename or current.strip()[:50]
            chunks.append(TextChunk(text=current.strip(), title=chunk_title, chunk_index=len(chunks)))
            overlap_text = current[-overlap:] if len(current) > overlap else ""
            current = overlap_text + para + "\n"
        else:
            current += para + "\n"

    if current.strip():
        chunk_title = title or filename or current.strip()[:50]
        chunks.append(TextChunk(text=current.strip(), title=chunk_title, chunk_index=len(chunks)))

    return chunks
